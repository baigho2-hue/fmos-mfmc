import openpyxl
from apps.utilisateurs.models import Utilisateur
import os
import sys

file_path = r"c:\Users\HP\Documents\fmos-mfmc\Liste Med6 2024-2025.xlsx"
log_path = "import_log_final.txt"

def log(msg):
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(msg + "\n")
        f.flush()
    print(msg)

def import_students():
    with open(log_path, "w", encoding="utf-8") as f:
        f.write("")
        
    log("=== Début de l'importation FINALE des étudiants Médecine 6 ===")
    
    try:
        if not os.path.exists(file_path):
            log(f"ERREUR: Fichier non trouvé: {file_path}")
            return

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active
        log(f"Fichier ouvert. Feuille: {sheet.title}")
        
        max_row = sheet.max_row
        log(f"Nombre de lignes estimé : {max_row}")
        
        rows = sheet.iter_rows(min_row=3, values_only=True)
        
        count_created = 0
        count_updated = 0
        count_errors = 0
        
        for i, row in enumerate(rows, 3):
            if not row[1]: 
                continue
            
            matricule = str(row[1]).strip()
            nom = str(row[2]).strip() if row[2] else ""
            prenom = str(row[3]).strip() if row[3] else ""
            
            if i % 50 == 0:
                log(f"Traitement ligne {i}...")
            
            email = f"{matricule.lower()}@fmos.ml"
            classe_nom = "Médecine 6"
            
            try:
                user, created = Utilisateur.objects.update_or_create(
                    username=matricule,
                    defaults={
                        'first_name': prenom,
                        'last_name': nom,
                        'email': email,
                        'type_utilisateur': 'etudiant',
                        'classe': classe_nom,
                        'is_active': True,
                        'email_verifie': True
                    }
                )
                
                if created:
                    user.set_password(matricule)
                    user.save()
                    count_created += 1
                else:
                    count_updated += 1
                    
            except Exception as e:
                count_errors += 1
                log(f"[ERREUR] Ligne {i} ({matricule}): {e}")
        
        log("\n=== Résumé de l'importation ===")
        log(f"Créés : {count_created}")
        log(f"Mis à jour : {count_updated}")
        log(f"Erreurs : {count_errors}")
        log("===============================")

    except Exception as e:
        log(f"Erreur critique lors de l'importation : {e}")
        import traceback
        log(traceback.format_exc())

import_students()
