"""
Fonctions utilitaires pour synchroniser les étudiants Med6 à partir d'un fichier Excel.
"""
from __future__ import annotations

from pathlib import Path
from typing import Callable, Dict, Optional

from django.conf import settings
from django.db import transaction

from apps.utilisateurs.models_med6 import EtudiantMed6, ListeMed6


LogFunction = Optional[Callable[[str, str], None]]


def _normalize_header(value) -> str:
    import unicodedata
    import re

    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = re.sub(r"\s+", " ", text)
    return text


def _detect_columns(headers):
    normalized = {_normalize_header(header): idx + 1 for idx, header in enumerate(headers)}

    def find_col(candidates, default):
        for candidate in candidates:
            if candidate in normalized:
                return normalized[candidate]
        return default

    matricule_col = find_col(
        ["matricule", "numero matricule", "numero"], default=1
    )
    prenom_col = find_col(
        ["prenom", "prénom", "prenoms", "prénoms"], default=2
    )
    nom_col = find_col(["nom", "noms"], default=3)
    return matricule_col, prenom_col, nom_col


def _resolve_file_path(fichier_source: str) -> Path:
    path = Path(fichier_source)
    if not path.is_absolute():
        path = Path(settings.BASE_DIR) / path
    return path


def sync_etudiants_from_excel(
    liste: ListeMed6,
    fichier_source: str,
    log: LogFunction = None,
) -> Dict[str, int]:
    """
    Crée ou met à jour les étudiants d'une liste Med6 à partir d'un fichier Excel.
    """
    if not fichier_source:
        raise ValueError("Aucun fichier source n'est associé à cette liste.")

    file_path = _resolve_file_path(fichier_source)
    if not file_path.exists():
        raise FileNotFoundError(f"Le fichier {file_path} est introuvable.")

    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("Le module openpyxl est requis pour cette opération.") from exc

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    matricule_col, prenom_col, nom_col = _detect_columns(headers)

    imported = 0
    updated = 0
    errors = 0

    def write_log(level: str, message: str):
        if log:
            log(level, message)

    with transaction.atomic():
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                matricule = row[matricule_col - 1].value if matricule_col <= len(row) else None
                prenom = row[prenom_col - 1].value if prenom_col <= len(row) else None
                nom = row[nom_col - 1].value if nom_col <= len(row) else None

                if not matricule and not prenom and not nom:
                    continue

                matricule = str(matricule).strip() if matricule else ""
                prenom = str(prenom).strip() if prenom else ""
                nom = str(nom).strip() if nom else ""

                if not matricule or not prenom or not nom:
                    write_log(
                        "warning",
                        f"Ligne {row_num}: données incomplètes (matricule={matricule}, prenom={prenom}, nom={nom})",
                    )
                    continue

                etudiant, created = EtudiantMed6.objects.get_or_create(
                    liste=liste,
                    matricule=matricule,
                    defaults={
                        "prenom": prenom,
                        "nom": nom,
                        "actif": True,
                    },
                )

                if created:
                    imported += 1
                    write_log("success", f"Importé: {prenom} {nom} ({matricule})")
                else:
                    etudiant.prenom = prenom
                    etudiant.nom = nom
                    etudiant.actif = True
                    etudiant.save()
                    updated += 1
                    write_log("success", f"Mis à jour: {prenom} {nom} ({matricule})")

            except Exception as exc:  # pylint: disable=broad-except
                errors += 1
                write_log("error", f"Erreur ligne {row_num}: {exc}")
                continue

        liste.nombre_etudiants = EtudiantMed6.objects.filter(liste=liste).count()
        liste.save(update_fields=["nombre_etudiants"])

    return {
        "imported": imported,
        "updated": updated,
        "errors": errors,
        "headers": headers,
        "matricule_col": matricule_col,
        "prenom_col": prenom_col,
        "nom_col": nom_col,
        "total": liste.nombre_etudiants,
    }


