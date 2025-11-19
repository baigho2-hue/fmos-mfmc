# apps/utilisateurs/management/commands/import_med6.py
"""
Commande Django pour importer les étudiants Med 6 depuis le fichier Excel
"""
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone
from apps.utilisateurs.models_med6 import EtudiantMed6, ListeMed6
import os
from pathlib import Path
from datetime import datetime


class Command(BaseCommand):
    help = 'Importe les étudiants de Médecine 6 depuis le fichier Excel Liste Med6 2024-2025.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='Liste Med6 2024-2025.xlsx',
            help='Chemin vers le fichier Excel à importer'
        )
        parser.add_argument(
            '--annee',
            type=str,
            help='Année universitaire (ex: 2024-2025). Si non spécifié, extrait du nom du fichier'
        )
        parser.add_argument(
            '--date-cloture',
            type=str,
            help='Date de clôture de l\'année universitaire (format: YYYY-MM-DD). Par défaut: 31 juillet de l\'année suivante'
        )
        parser.add_argument(
            '--desactiver-anciennes',
            action='store_true',
            help='Désactiver automatiquement les anciennes listes'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        annee_universitaire = options.get('annee')
        date_cloture_str = options.get('date_cloture')
        desactiver_anciennes = options.get('desactiver_anciennes', False)
        
        # Vérifier si le fichier existe
        if not os.path.exists(file_path):
            self.stdout.write(
                self.style.ERROR(f'Le fichier {file_path} n\'existe pas.')
            )
            return
        
        # Essayer d'importer openpyxl
        try:
            import openpyxl
        except ImportError:
            self.stdout.write(
                self.style.ERROR('Le module openpyxl n\'est pas installé. Installez-le avec: pip install openpyxl')
            )
            return
        
        # Déterminer l'année universitaire
        if not annee_universitaire:
            # Essayer d'extraire depuis le nom du fichier
            filename = os.path.basename(file_path)
            import re
            match = re.search(r'(\d{4})-(\d{4})', filename)
            if match:
                annee_universitaire = f"{match.group(1)}-{match.group(2)}"
            else:
                # Par défaut, utiliser l'année actuelle
                current_year = timezone.now().year
                annee_universitaire = f"{current_year}-{current_year + 1}"
        
        # Déterminer la date de clôture
        if date_cloture_str:
            try:
                date_cloture = datetime.strptime(date_cloture_str, '%Y-%m-%d').date()
            except ValueError:
                self.stdout.write(
                    self.style.ERROR(f'Format de date invalide. Utilisez YYYY-MM-DD')
                )
                return
        else:
            # Par défaut: 31 juillet de l'année suivante
            annee_suivante = int(annee_universitaire.split('-')[1])
            date_cloture = datetime(annee_suivante, 7, 31).date()
        
        # Vérifier si une liste existe déjà pour cette année
        try:
            liste_existante = ListeMed6.objects.get(annee_universitaire=annee_universitaire)
            self.stdout.write(
                self.style.WARNING(f'Une liste existe déjà pour l\'année {annee_universitaire}.')
            )
            reponse = input('Voulez-vous la remplacer ? (oui/non): ')
            if reponse.lower() not in ['oui', 'o', 'yes', 'y']:
                self.stdout.write(self.style.WARNING('Import annulé.'))
                return
            # Désactiver l'ancienne liste
            liste_existante.active = False
            liste_existante.save()
        except ListeMed6.DoesNotExist:
            pass
        
        # Désactiver les anciennes listes si demandé
        if desactiver_anciennes:
            ListeMed6.objects.filter(active=True).update(active=False)
            self.stdout.write(self.style.SUCCESS('Anciennes listes désactivées.'))
        
        # Créer ou mettre à jour la liste
        liste, created = ListeMed6.objects.get_or_create(
            annee_universitaire=annee_universitaire,
            defaults={
                'date_cloture': date_cloture,
                'fichier_source': os.path.basename(file_path),
                'active': True
            }
        )
        
        if not created:
            liste.date_cloture = date_cloture
            liste.fichier_source = os.path.basename(file_path)
            liste.active = True
            liste.save()
        
        self.stdout.write(
            self.style.SUCCESS(f'Liste {annee_universitaire} créée/mise à jour. Date de clôture: {date_cloture}')
        )
        
        try:
            # Charger le fichier Excel
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Lire les en-têtes (première ligne)
            headers = [cell.value for cell in ws[1]]
            self.stdout.write(f'Colonnes trouvées: {headers}')

            def normalize_header(value):
                import unicodedata
                import re
                if value is None:
                    return ""
                value = str(value).strip().lower()
                value = unicodedata.normalize('NFD', value)
                value = ''.join(c for c in value if unicodedata.category(c) != 'Mn')
                value = re.sub(r'\s+', ' ', value)
                return value

            normalized_headers = {normalize_header(header): idx + 1 for idx, header in enumerate(headers)}

            def find_col(candidates, default=None):
                for candidate in candidates:
                    if candidate in normalized_headers:
                        return normalized_headers[candidate]
                return default

            # Essayer de déterminer dynamiquement les colonnes à partir des en-têtes
            matricule_col = find_col(
                ['matricule', 'numero matricule', 'numero'],
                default=1  # Par défaut, première colonne
            )
            prenom_col = find_col(
                ['prenom', 'prénom', 'prenoms', 'prénoms'],
                default=2  # Par défaut, deuxième colonne
            )
            nom_col = find_col(
                ['nom', 'noms'],
                default=3  # Par défaut, troisième colonne
            )

            self.stdout.write(
                self.style.SUCCESS(
                    f'Colonnes utilisées -> Matricule: {matricule_col}, Prénom: {prenom_col}, Nom: {nom_col}'
                )
            )
            
            # Parcourir les lignes (en commençant à la ligne 2 pour ignorer l'en-tête)
            imported = 0
            updated = 0
            errors = 0
            
            with transaction.atomic():
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                    try:
                        # Récupérer les valeurs des colonnes
                        matricule = row[matricule_col - 1].value if matricule_col <= len(row) else None
                        prenom = row[prenom_col - 1].value if prenom_col <= len(row) else None
                        nom = row[nom_col - 1].value if nom_col <= len(row) else None
                        
                        # Ignorer les lignes vides
                        if not matricule and not prenom and not nom:
                            continue
                        
                        # Nettoyer les valeurs
                        matricule = str(matricule).strip() if matricule else ""
                        prenom = str(prenom).strip() if prenom else ""
                        nom = str(nom).strip() if nom else ""
                        
                        if not matricule or not prenom or not nom:
                            self.stdout.write(
                                self.style.WARNING(f'Ligne {row_num}: Données incomplètes (matricule: {matricule}, prenom: {prenom}, nom: {nom})')
                            )
                            continue
                        
                        # Créer ou mettre à jour l'étudiant dans cette liste
                        etudiant, created = EtudiantMed6.objects.get_or_create(
                            liste=liste,
                            matricule=matricule,
                            defaults={
                                'prenom': prenom,
                                'nom': nom,
                                'actif': True
                            }
                        )
                        
                        if created:
                            imported += 1
                            self.stdout.write(
                                self.style.SUCCESS(f'Importé: {prenom} {nom} ({matricule})')
                            )
                        else:
                            # Mettre à jour les informations
                            etudiant.prenom = prenom
                            etudiant.nom = nom
                            etudiant.actif = True
                            etudiant.save()
                            updated += 1
                            self.stdout.write(
                                self.style.SUCCESS(f'Mis à jour: {prenom} {nom} ({matricule})')
                            )
                    
                    except Exception as e:
                        errors += 1
                        self.stdout.write(
                            self.style.ERROR(f'Erreur ligne {row_num}: {str(e)}')
                        )
                        continue
            
            # Mettre à jour le nombre d'étudiants dans la liste
            liste.nombre_etudiants = EtudiantMed6.objects.filter(liste=liste).count()
            liste.save()

            # Afficher les informations sur l'expiration
            jours_avant_expiration = liste.jours_avant_expiration()
            if jours_avant_expiration is not None:
                if jours_avant_expiration < 0:
                    self.stdout.write(
                        self.style.WARNING(
                            f'\nATTENTION: Cette liste est expirée depuis {abs(jours_avant_expiration)} jours!'
                        )
                    )
                else:
                    self.stdout.write(
                        self.style.SUCCESS(
                            f'\nCette liste sera valide pendant encore {jours_avant_expiration} jours'
                        )
                    )

            self.stdout.write(
                self.style.SUCCESS(
                    f'\nImport terminé: {imported} importés, {updated} mis à jour, {errors} erreurs'
                )
            )
            self.stdout.write(
                self.style.SUCCESS(
                    f'Total étudiants dans la liste: {liste.nombre_etudiants}'
                )
            )

        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'Erreur lors de l\'import: {str(e)}')
            )

